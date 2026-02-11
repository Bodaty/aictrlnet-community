"""File processing node — reads StagedFile, extracts structured data.

Supports PDF (pdfplumber), Excel (openpyxl), and CSV (stdlib).
Output: structured dict to next node in the workflow.
"""

import csv
import io
import logging
from datetime import datetime
from typing import Any, Dict

from ..base_node import BaseNode
from ..models import NodeConfig, NodeExecutionResult, NodeStatus
from events.event_bus import event_bus

logger = logging.getLogger(__name__)


class FileProcessNode(BaseNode):
    """Node for extracting structured data from uploaded files.

    Reads a StagedFile by file_id, detects format, and extracts content
    into a structured dict that downstream nodes can consume.

    Supported formats:
    - PDF  → page-by-page text extraction via pdfplumber
    - XLSX → sheet-by-sheet rows via openpyxl
    - CSV  → list of row dicts via csv.DictReader
    - TXT  → raw text content
    """

    async def execute(self, input_data: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """Execute file processing."""
        start_time = datetime.utcnow()

        try:
            # Get file reference — from node config or input data
            file_id = self.config.parameters.get("file_id") or input_data.get("file_id")
            file_path = self.config.parameters.get("file_path") or input_data.get("file_path")
            content_type = self.config.parameters.get("content_type") or input_data.get("content_type", "")

            if not file_path and not file_id:
                raise ValueError("Either file_id or file_path is required")

            # If we have file_id but no path, resolve from staged files
            if file_id and not file_path:
                file_path = input_data.get("storage_path") or f"/tmp/aictrlnet/staged_files/{file_id}"

            # Read raw bytes
            with open(file_path, "rb") as f:
                raw_bytes = f.read()

            # Detect content type from extension if not provided
            if not content_type:
                content_type = self._detect_content_type(file_path)

            # Route to format-specific processor
            if "pdf" in content_type:
                extracted = self._process_pdf(raw_bytes)
            elif "spreadsheet" in content_type or "excel" in content_type or file_path.endswith((".xlsx", ".xls")):
                extracted = self._process_excel(raw_bytes)
            elif "csv" in content_type or file_path.endswith(".csv"):
                extracted = self._process_csv(raw_bytes)
            else:
                # Fallback: treat as plain text
                extracted = self._process_text(raw_bytes)

            duration_ms = (datetime.utcnow() - start_time).total_seconds() * 1000

            await event_bus.publish(
                "node.executed",
                {
                    "node_id": self.config.id,
                    "node_type": "fileProcess",
                    "content_type": content_type,
                    "duration_ms": duration_ms,
                },
            )

            return {
                "file_id": file_id,
                "content_type": content_type,
                "extracted": extracted,
                "row_count": extracted.get("row_count", 0),
                "page_count": extracted.get("page_count", 0),
            }

        except Exception as e:
            logger.error(f"FileProcessNode {self.config.id} failed: {e}")
            raise

    def _process_pdf(self, raw_bytes: bytes) -> Dict[str, Any]:
        """Extract text from PDF using pdfplumber."""
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber is required for PDF processing. Install with: pip install pdfplumber")

        pages = []
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                tables = page.extract_tables() or []
                pages.append({
                    "page_number": i + 1,
                    "text": text,
                    "tables": tables,
                    "width": float(page.width),
                    "height": float(page.height),
                })

        # Combine all text for convenience
        full_text = "\n\n".join(p["text"] for p in pages if p["text"])

        return {
            "type": "pdf",
            "page_count": len(pages),
            "pages": pages,
            "full_text": full_text,
            "total_tables": sum(len(p["tables"]) for p in pages),
        }

    def _process_excel(self, raw_bytes: bytes) -> Dict[str, Any]:
        """Extract data from Excel using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel processing. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        sheets = {}
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if i == 0:
                    headers = row_data
                else:
                    if headers:
                        rows.append(dict(zip(headers, row_data)))
                    else:
                        rows.append(row_data)

            sheets[sheet_name] = {
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
            }
            total_rows += len(rows)

        wb.close()

        return {
            "type": "excel",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "row_count": total_rows,
        }

    def _process_csv(self, raw_bytes: bytes) -> Dict[str, Any]:
        """Extract data from CSV using stdlib."""
        text = raw_bytes.decode("utf-8-sig")  # Handle BOM
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

        return {
            "type": "csv",
            "headers": reader.fieldnames or [],
            "rows": rows,
            "row_count": len(rows),
        }

    def _process_text(self, raw_bytes: bytes) -> Dict[str, Any]:
        """Process plain text file."""
        text = raw_bytes.decode("utf-8", errors="replace")
        lines = text.splitlines()

        return {
            "type": "text",
            "text": text,
            "line_count": len(lines),
            "char_count": len(text),
        }

    def _detect_content_type(self, file_path: str) -> str:
        """Detect content type from file extension."""
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
        mapping = {
            "pdf": "application/pdf",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xls": "application/vnd.ms-excel",
            "csv": "text/csv",
            "txt": "text/plain",
            "json": "application/json",
        }
        return mapping.get(ext, "application/octet-stream")

    def validate_config(self) -> bool:
        """Validate node configuration."""
        # file_id or file_path can come from input_data at runtime
        return True
